import re
import pandas as pd
import json
import os
import numpy as np
from datetime import datetime
import cv2
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage


class LogCore:
    THUMBNAIL_MAX_DIM = 85
    CONFIG_FILE = 'ui_config.json'

    ID_PATTERN = r"(?:零件?唯一编号|唯一编号|unique id|partCodeUnique|partcodeunique)"

    REGEX_TEXT_LOGS = {
        '放置报工完成': re.compile(r"(?P<station>\d+)号台面，(?P<mat_code>[\w-]+)零件,唯一编码(?P<seq>\w+)放置报工完成"),
        '报工前置ID': re.compile(r"报工台号:\d+,零件唯一编号:(?P<seq>\w+)"),
        '零件入框成功': re.compile(r"【零件入框】成功。零件：(?P<seq>\w+)"),
        '小件一次分拣开始': re.compile(rf"【小件一次分拣开始】.*?{ID_PATTERN}[:：]?\s*(?P<seq>\w+).*?零件号(?P<mat_code>[\w-]+).*?任务号[:：]?\s*(?P<task_no>[\w]+)"),
        '小件一次分拣完成': re.compile(rf"【小件一次分拣完成】.*?{ID_PATTERN}[:：]?\s*(?P<seq>\w+).*?零件号(?P<mat_code>[\w-]+).*?任务号[:：]?\s*(?P<task_no>[\w]+)"),
        '小件二次分拣开始': re.compile(rf"【小件二次分拣机器人分拣开始】.*?{ID_PATTERN}[:：]?\s*(?P<seq>\w+)(?:.*?任务号[:：]?\s*(?P<task_no>[\w]+))?"),
        '小件二次分拣完成': re.compile(rf"【小件二次分拣机器人分拣完成】.*?{ID_PATTERN}[:：]?\s*(?P<seq>\w+)(?:.*?任务号[:：]?\s*(?P<task_no>[\w]+))?"),
        '大件桁架分拣开始': re.compile(rf"【桁架分拣开始】.*?{ID_PATTERN}[:：]?\s*(?P<seq>\w+)(?:.*?任务号[:：]?\s*(?P<task_no>[\w]+))?"),
        '大件桁架分拣完成': re.compile(rf"【桁架分拣完成】.*?{ID_PATTERN}[:：]?\s*(?P<seq>\w+)(?:.*?任务号[:：]?\s*(?P<task_no>[\w]+))?"),
        '暗室流转判定': re.compile(r"\[(?P<station_no>AS\d+)\].*?零件(?P<mat_code>[\w-]+)流向为:(?P<flow>\w+)"),
        '具体异常': re.compile(r"零件(?P<seq>\w+)(?P<err_msg>异常|抓取失败|掉落)"),
    }

    REGEX_JSON_LOGS = {
        '满框置换请求': re.compile(r"【满框置换请求】.*?(?P<json_data>\{.*\})"),
        '坡口倒棱报工': re.compile(r"坡口倒棱任务完成报工.*?(?P<json_data>\{.*\})"),
        '零件分拣报工': re.compile(r"【零件报工】.*?(?P<json_data>\{.*\})"),
        '倒棱任务下发': re.compile(r"【向倒棱软件下发任务】.*?(?P<json_data>\{.*\})"),
    }

    REGEX_STEEL_SIGNAL = re.compile(r"收到钢板到位信号：(?P<json_data>\{.*\})")
    REGEX_PARAM_LINE = re.compile(r"参数：(?P<json_data>\{.*\})")
    REGEX_RESULT_LINE = re.compile(r"结果：(?P<json_data>\{.*\})")
    BASE_TIME_PATTERN = re.compile(r"^(?P<time>\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\.\d{3})")

    def __init__(self):
        self.nesting_index = {}
        self.log_root = ""
        self.nesting_root = ""
        self.load_config()

    def load_config(self):
        if os.path.exists(self.CONFIG_FILE):
            try:
                with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.log_root = config.get('log_root', '')
                    self.nesting_root = config.get('nesting_root', '')
                    if self.nesting_root and os.path.exists(self.nesting_root):
                        self.build_physical_nesting_index(self.nesting_root)
            except: pass

    def save_config(self, log_root, nesting_root):
        self.log_root = log_root
        self.nesting_root = nesting_root
        with open(self.CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump({'log_root': log_root, 'nesting_root': nesting_root}, f, ensure_ascii=False, indent=4)
        if nesting_root and os.path.exists(nesting_root):
            self.build_physical_nesting_index(nesting_root)

    def build_physical_nesting_index(self, nesting_root):
        print(f"正在盘点物理套料图库: {nesting_root} ...")
        self.nesting_index.clear()
        for date_dir in os.listdir(nesting_root):
            date_path = os.path.join(nesting_root, date_dir)
            if not os.path.isdir(date_path): continue

            for task_no in os.listdir(date_path):
                task_path = os.path.join(date_path, task_no)
                if not os.path.isdir(task_path): continue

                for file_name in os.listdir(task_path):
                    m = re.match(r"^(\d+)_([A-Za-z0-9]+)(_.*)?\.(?:dnf|png|jpg)$", file_name, re.IGNORECASE)
                    if m:
                        sort_seq = m.group(1)
                        part_bgr = m.group(2)
                        suffix = m.group(3)
                        current_priority = 0 if not suffix else 1

                        if task_no not in self.nesting_index:
                            self.nesting_index[task_no] = {}

                        current_record = self.nesting_index[task_no].get(sort_seq, {})
                        if not current_record or current_priority < current_record.get("priority", 99):
                            self.nesting_index[task_no][sort_seq] = {
                                "part_bgr": part_bgr,
                                "file_path": os.path.join(task_path, file_name),
                                "priority": current_priority
                            }

    @staticmethod
    def parse_json_safe(json_str):
        try: return json.loads(json_str)
        except: return None

    @staticmethod
    def calculate_duration(history_list):
        if not history_list: return 0.0, "", ""
        try:
            times = []
            ptn = re.compile(r"\[(\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\.\d{3})\]")
            for h in history_list:
                m = ptn.search(h)
                if m: times.append(datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S.%f"))
            if len(times) >= 2:
                times.sort()
                duration = (times[-1] - times[0]).total_seconds()
                return round(duration / 60, 2), str(times[0]), str(times[-1])
        except: pass
        return 0.0, "", ""

    def process_folder(self, folder_path):
        print(f"\n   -> 正在分析分拣数据: {os.path.basename(folder_path)} ...")
        parts_db = {}
        temp_station_map = {}
        current_time = "Unknown Time"
        pending_report_id = None
        current_context = None
        global_task_no = None

        if not os.path.exists(folder_path): return None, "路径不存在"
        if os.path.isfile(folder_path): folder_path = os.path.dirname(folder_path)

        files = [f for f in os.listdir(folder_path) if f.endswith(('.log', '.txt'))]
        if not files: return None, "无日志文件"

        def get_log_index(filename):
            m = re.search(r'\.(\d+)\.(?:log|txt)$', filename)
            return int(m.group(1)) if m else 0

        files.sort(key=get_log_index)

        for file_name in files:
            file_path = os.path.join(folder_path, file_name)
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    for line in f:
                        t_match = self.BASE_TIME_PATTERN.match(line)
                        if t_match: current_time = t_match.group('time')

                        if "===>请求【虚拟托盘码盘】开始" in line: current_context = "VIRTUAL_PACKING"
                        elif "===>请求【码盘调度】开始" in line: current_context = "STACKING_SCHEDULING"
                        elif "===>请求" in line: current_context = None

                        m_signal = self.REGEX_STEEL_SIGNAL.search(line)
                        if m_signal:
                            data = self.parse_json_safe(m_signal.group('json_data'))
                            if data and 'data' in data: global_task_no = data['data'].get('task_no')

                        processed_text = False
                        for action_name, pattern in self.REGEX_TEXT_LOGS.items():
                            match = pattern.search(line)
                            if match:
                                groups = match.groupdict()
                                u_id = groups.get('seq')
                                if u_id:
                                    if action_name == '报工前置ID':
                                        pending_report_id = u_id
                                        continue
                                    if u_id not in parts_db:
                                        parts_db[u_id] = {'history': [], 'type_hints': set(), 'part_code': groups.get('mat_code', ''), 'has_scheduled': False, 'specific_error': None, 'station_no': '', 'frame_code': ''}
                                    if groups.get('mat_code'): parts_db[u_id]['part_code'] = groups.get('mat_code')
                                    if groups.get('task_no'): parts_db[u_id]['task_no'] = groups.get('task_no')
                                    if groups.get('station'): parts_db[u_id]['station_no'] = groups.get('station')

                                    if action_name == '具体异常' and groups.get('err_msg'):
                                        parts_db[u_id]['specific_error'] = groups.get('err_msg')
                                        parts_db[u_id]['history'].append(f"[{current_time}] 异常: {groups.get('err_msg')}")
                                    else:
                                        parts_db[u_id]['history'].append(f"[{current_time}] {action_name}")

                                    if '小件' in action_name: parts_db[u_id]['type_hints'].add('小件')
                                    elif '大件' in action_name: parts_db[u_id]['type_hints'].add('大件')
                                    processed_text = True
                                    break
                        if processed_text: continue

                        if current_context == "VIRTUAL_PACKING":
                            m_param = self.REGEX_PARAM_LINE.search(line)
                            if m_param:
                                data = self.parse_json_safe(m_param.group('json_data'))
                                if data and 'partList' in data:
                                    parent_task = data.get('taskCode') or data.get('task_no') or global_task_no
                                    for mat in data['partList']:
                                        u_id = mat.get('partCodeUnique') or mat.get('seq')
                                        if not u_id: continue
                                        if u_id not in parts_db:
                                            parts_db[u_id] = {'history': [], 'type_hints': set(), 'part_code': mat.get('partCode'), 'has_scheduled': False, 'specific_error': None, 'station_no': '', 'frame_code': ''}
                                        if parent_task: parts_db[u_id]['task_no'] = parent_task
                                        if mat.get('partCode'): parts_db[u_id]['part_code'] = mat.get('partCode')
                                        p_type = str(mat.get('partType', ''))
                                        if p_type == '4': parts_db[u_id]['type_hints'].add('大件')
                                        elif p_type == '2': parts_db[u_id]['type_hints'].add('小件')
                                        parts_db[u_id]['history'].append(f"[{current_time}] 虚拟托盘请求")

                        elif current_context == "STACKING_SCHEDULING":
                            for regex in [self.REGEX_PARAM_LINE, self.REGEX_RESULT_LINE]:
                                m = regex.search(line)
                                if m:
                                    data = self.parse_json_safe(m.group('json_data'))
                                    part_list = data.get('partList', []) if data else []
                                    if data and 'data' in data and 'partList' in data['data']: part_list = data['data']['partList']
                                    for mat in part_list:
                                        u_id = mat.get('partCodeUnique')
                                        if not u_id: continue
                                        if u_id not in parts_db:
                                            parts_db[u_id] = {'history': [], 'type_hints': set(), 'part_code': mat.get('partCode'), 'has_scheduled': False, 'specific_error': None, 'station_no': '', 'frame_code': ''}
                                        if mat.get('partCode'): parts_db[u_id]['part_code'] = mat.get('partCode')
                                        if regex == self.REGEX_RESULT_LINE:
                                            parts_db[u_id]['has_scheduled'] = True
                                            if mat.get('frameCode'): parts_db[u_id]['frame_code'] = mat.get('frameCode')
                                            parts_db[u_id]['history'].append(f"[{current_time}] 码盘调度完成")

                        m_report = self.REGEX_JSON_LOGS['零件分拣报工'].search(line)
                        if m_report:
                            data = self.parse_json_safe(m_report.group('json_data'))
                            if data and 'data' in data:
                                d = data['data']
                                u_id = d.get('seq') or d.get('partCodeUnique') or pending_report_id
                                pending_report_id = None

                                station = d.get('station_no')
                                frame = d.get('pallet_code') or d.get('frame_code')
                                mat_code = d.get('mat_code')
                                if mat_code and (station or frame):
                                    temp_station_map[mat_code] = {'station': station, 'frame': frame}

                                if u_id:
                                    if u_id not in parts_db: parts_db[u_id] = {'history': [], 'type_hints': set(), 'part_code': '', 'has_scheduled': False, 'specific_error': None, 'station_no': '', 'frame_code': ''}
                                    if d.get('mat_code'): parts_db[u_id]['part_code'] = d.get('mat_code')
                                    if d.get('task_no'): parts_db[u_id]['task_no'] = d.get('task_no')
                                    if station: parts_db[u_id]['station_no'] = station
                                    if frame: parts_db[u_id]['frame_code'] = frame
                                    parts_db[u_id]['history'].append(f"[{current_time}] 分拣报工完成")
            except Exception as e:
                print(f"Error reading file {file_name}: {e}")
                continue

        export_list = []
        FINISH_KEYWORDS = ["分拣报工", "零件入框", "坡口倒棱", "倒棱任务", "放置报工完成", "分拣完成", "桁架分拣完成"]

        thumb_cache_dir = os.path.join(folder_path, ".ui_thumbs_cache")
        if not os.path.exists(thumb_cache_dir): os.makedirs(thumb_cache_dir)

        color_map = {
            '🟢 正常': (0, 255, 0),
            '🟡 警戒': (0, 255, 255),
            '🔴': (0, 0, 255),
            '🤍 不参与分拣': (180, 180, 180)
        }

        def get_event_times(history_list, event_keyword):
            times = []
            ptn = re.compile(r"\[(\d{4}-\d{2}-\d{2}\s\d{2}:\d{2}:\d{2}\.\d{3})\]")
            for h in history_list:
                if event_keyword in h:
                    m = ptn.search(h)
                    if m: times.append(datetime.strptime(m.group(1), "%Y-%m-%d %H:%M:%S.%f"))
            return sorted(times)

        for u_id, info in parts_db.items():
            if re.search(r'[\u4e00-\u9fa5]', str(u_id)): continue
            if re.search(r'[\u4e00-\u9fa5]', str(info.get('part_code', ''))): continue

            info['history'].sort()

            mat = info.get('part_code')
            if mat and mat in temp_station_map:
                if not info.get('station_no'): info['station_no'] = temp_station_map[mat]['station']
                if not info.get('frame_code'): info['frame_code'] = temp_station_map[mat]['frame']

            duration_min, t_start, t_end = self.calculate_duration(info['history'])
            hints = info.get('type_hints', set())
            p_type = '大件' if '大件' in hints else ('小件' if '小件' in hints else '未知')
            is_finished = any(k in str(info['history']) for k in FINISH_KEYWORDS)

            info.update({'duration_min': duration_min, 'is_finished': is_finished, 'p_type': p_type, 't_start': t_start, 't_end': t_end})

            t_pri_start, t_pri_end = get_event_times(info['history'], '小件一次分拣开始'), get_event_times(info['history'], '小件一次分拣完成')
            dur_primary = round((t_pri_end[-1] - t_pri_start[0]).total_seconds() / 60, 2) if t_pri_start and t_pri_end else -1.0

            t_sec_start, t_sec_end = get_event_times(info['history'], '小件二次分拣开始'), get_event_times(info['history'], '小件二次分拣完成')
            dur_secondary = round((t_sec_end[-1] - t_sec_start[0]).total_seconds() / 60, 2) if t_sec_start and t_sec_end else -1.0

            t_truss_start, t_truss_end = get_event_times(info['history'], '大件桁架分拣开始'), get_event_times(info['history'], '大件桁架分拣完成')
            dur_truss = round((t_truss_end[-1] - t_truss_start[0]).total_seconds() / 60, 2) if t_truss_start and t_truss_end else -1.0

            t_pallet = get_event_times(info['history'], '码盘调度完成')
            dur_pallet = -1.0
            if len(t_pallet) == 1: dur_pallet = 1.0
            elif len(t_pallet) > 1: dur_pallet = round((t_pallet[-1] - t_pallet[0]).total_seconds() / 60, 2)

            def get_severity(dur, limit_n, limit_w):
                if dur < 0: return 0
                if dur <= limit_n: return 0
                elif dur <= limit_w: return 1
                else: return 2

            sev_pri = get_severity(dur_primary, 1.0, 30.0)
            sev_sec = get_severity(dur_secondary, 1.0, 30.0)
            sev_tru = get_severity(dur_truss, 15.0, 40.0)
            sev_pal = get_severity(dur_pallet, 1.0, 30.0)

            step_sevs = [
                (sev_pri, "一次分拣"),
                (sev_sec, "二次分拣"),
                (sev_tru, "桁架分拣"),
                (sev_pal, "码盘调度")
            ]
            max_sev = max(s[0] for s in step_sevs)
            faulty_steps = [s[1] for s in step_sevs if s[0] == max_sev]
            faulty_name = "+".join(faulty_steps) if faulty_steps else ""

            status = "🟢 正常"
            if info.get('specific_error'):
                status = f"🔴 异常:{info['specific_error']}"
            elif info['duration_min'] == 0:
                status = "🤍 不参与分拣"
            elif info['has_scheduled'] and not info['is_finished']:
                status = "🔴 异常(丢失)"
            elif info['duration_min'] > 0 or info['is_finished']:
                if max_sev == 2: status = f"🔴 异常({faulty_name}超时)"
                elif max_sev == 1: status = f"🟡 警戒({faulty_name}超时)"
                else: status = "🟢 正常"
            else:
                status = "🤍 不参与分拣"

            part_bgr = ""
            full_image_path = ""
            ui_thumb_path = ""
            task_no = info.get('task_no', '')
            u_id_str = str(u_id)

            if task_no and len(u_id_str) > 4 and u_id_str.isdigit():
                real_sort_seq = u_id_str[:-4]
                if task_no in self.nesting_index and real_sort_seq in self.nesting_index[task_no]:
                    target_data = self.nesting_index[task_no][real_sort_seq]
                    part_bgr = target_data["part_bgr"]
                    full_image_path = target_data["file_path"]

                    if full_image_path and os.path.exists(full_image_path) and full_image_path.lower().endswith('.png'):
                        try:
                            img_color = cv2.imread(full_image_path, cv2.IMREAD_COLOR)
                            if img_color is not None:
                                target_bgr = (255, 255, 255)
                                for key, bgr in color_map.items():
                                    if key in status: target_bgr = bgr; break

                                white_mask_bool = (img_color[:, :, 2] > 200) & (img_color[:, :, 1] > 200) & (img_color[:, :, 0] > 200)
                                base_mask = np.zeros(img_color.shape[:2], dtype=np.uint8)
                                base_mask[white_mask_bool] = 255

                                kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
                                thick_mask = cv2.dilate(base_mask, kernel, iterations=5)

                                img_final = np.zeros_like(img_color)
                                img_final[thick_mask == 255] = target_bgr

                                h, w = img_final.shape[:2]
                                scale = min(self.THUMBNAIL_MAX_DIM / w, self.THUMBNAIL_MAX_DIM / h)
                                new_w, new_h = max(1, int(w * scale)), max(1, int(h * scale))
                                img_resized = cv2.resize(img_final, (new_w, new_h), interpolation=cv2.INTER_AREA)

                                ui_thumb_path = os.path.join(thumb_cache_dir, f"{u_id}.png")
                                cv2.imwrite(ui_thumb_path, img_resized)
                        except Exception as e:
                            pass

            export_list.append({
                '唯一编号 (Unique ID)': u_id,
                '零件号 (Part Code)': info.get('part_code', ''),
                '套料图元号': part_bgr,
                '套料图物理路径': full_image_path,
                '零件类型': info['p_type'],
                '状态': status,
                '一次分拣耗时': dur_primary,
                '二次分拣耗时': dur_secondary,
                '桁架分拣耗时': dur_truss,
                '码盘调度耗时': dur_pallet,
                'UI微缩图': ui_thumb_path,
                '总耗时(分钟)': info['duration_min'],
                '任务号': task_no,
                '料框号': info.get('frame_code', ''),
                '最新工位': info.get('station_no', ''),
                '开始时间': info['t_start'],
                '结束时间': info['t_end'],
                '完整流程追踪': "\n".join(info['history'])
            })

        df = pd.DataFrame(export_list)
        if df.empty: return None, "未提取到有效数据"
        return df, "Success"

    def export_excel_with_images(self, df, save_path):
        print(f"\n正在生成带图报表: {os.path.basename(save_path)} ...")
        try:
            export_df = df.drop(columns=['UI微缩图']) if 'UI微缩图' in df.columns else df
            export_df.to_excel(save_path, index=False)

            wb = load_workbook(save_path)
            ws = wb.active

            col_to_shift = -1
            for cell in ws[1]:
                if cell.value == '完整流程追踪':
                    col_to_shift = cell.column
                    break

            if col_to_shift == -1: col_to_shift = ws.max_column + 1

            ws.insert_cols(col_to_shift)
            img_col_letter = ws.cell(row=1, column=col_to_shift).column_letter
            ws[f'{img_col_letter}1'] = '图形可视化'
            ws.column_dimensions[img_col_letter].width = 13

            for row_idx, row in df.iterrows():
                excel_row = row_idx + 2
                thumb_path = row.get('UI微缩图')
                if thumb_path and os.path.exists(thumb_path):
                    try:
                        xl_img = ExcelImage(thumb_path)
                        ws.add_image(xl_img, f"{img_col_letter}{excel_row}")
                        ws.row_dimensions[excel_row].height = 68
                    except Exception:
                        pass

            wb.save(save_path)
            print("   Excel 导出成功！")
            return True
        except Exception as e:
            print(f"操作 Excel 失败: {e}")
            return False
